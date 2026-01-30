#!/usr/bin/env python3
"""
Marketing Analytics Feed - Unified Data Aggregation
Modern Mangal Internal Operations

Consolidates all marketing and sales data into a single timeline:
- Shopify sales data
- Website traffic (Google Analytics)
- Social media influencer posts
- Ad campaign launches and creative changes

This creates a unified view to correlate marketing activities with sales performance.

Created: 2025-12-28
"""

import os
import sys
import json
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Import API connectors
from api_connectors import (
    InstagramConnector,
    TikTokConnector,
    PinterestConnector,
    MetaAdsConnector,
    GoogleAdsConnector,
    GoogleAnalytics4Connector,
    ShopifyAnalyticsConnector
)

# ============================================================================
# CONFIGURATION
# ============================================================================

# Path structure for marketing-feed-analytics repository
SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent  # marketing-feed-analytics/
SECRETS_DIR = REPO_ROOT.parent.parent / "secrets"  # /srv/secrets
SRV_DATA_DIR = REPO_ROOT.parent.parent / "data"  # /srv/data
ORDERS_DIR = SRV_DATA_DIR / "orders"
SHOPIFY_DATA_DIR = SRV_DATA_DIR / "shopify"

# Repository-local directories
DATA_DIR = REPO_ROOT / "data"
CONFIG_DIR = REPO_ROOT / "config"

# Output configuration
OUTPUT_FILE = DATA_DIR / 'Marketing_Analytics_Feed.xlsx'
CONFIG_FILE = CONFIG_DIR / 'marketing_config.json'

# Ensure directories exist
DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================================
# DATA CONNECTORS
# ============================================================================

class ShopifyDataConnector:
    """Extract sales and traffic data from Shopify"""
    
    def __init__(self, credentials_file: Path):
        self.credentials = self._load_credentials(credentials_file)
        self.shop_domain = 'modernmangal.myshopify.com'
        self.api_version = '2024-10'
        
    def _load_credentials(self, file_path: Path) -> Dict[str, str]:
        """Load Shopify API credentials"""
        creds = {}
        if file_path.exists():
            with open(file_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if '=' in line and not line.startswith('#'):
                        key, value = line.split('=', 1)
                        creds[key.strip()] = value.strip()
        return creds
    
    def get_daily_sales(self, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """
        Get daily sales aggregated from Shopify orders
        Returns: DataFrame with columns [date, order_count, total_revenue, avg_order_value]
        """
        # Try to load from existing order tracker first
        tracker_file = ORDERS_DIR / 'Shopify_Master_Order_Tracker.xlsx'
        
        if tracker_file.exists():
            print(f"Loading sales data from {tracker_file}")
            df = pd.read_excel(tracker_file, sheet_name='Master Order Dump')
            
            # Convert created_at to datetime (handle mixed timezones)
            df['Created At'] = pd.to_datetime(df['Created At'], utc=True, errors='coerce')
            df = df.dropna(subset=['Created At'])  # Remove rows with invalid dates
            df['date'] = df['Created At'].dt.date
            
            # Filter date range
            mask = (df['date'] >= start_date.date()) & (df['date'] <= end_date.date())
            df_filtered = df[mask]
            
            # Aggregate by day
            daily_sales = df_filtered.groupby('date').agg({
                'Order ID': 'count',
                'Total Price': 'sum'
            }).reset_index()
            
            daily_sales.columns = ['date', 'order_count', 'total_revenue']
            daily_sales['avg_order_value'] = daily_sales['total_revenue'] / daily_sales['order_count']
            
            return daily_sales
        else:
            print("Warning: Order tracker not found. Run comprehensive_order_tracker.py first.")
            return pd.DataFrame(columns=['date', 'order_count', 'total_revenue', 'avg_order_value'])
    
    def get_shopify_analytics(self, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """
        Get Shopify Analytics data (sessions, page views)
        Note: Requires Shopify Analytics API or third-party integration
        """
        # Placeholder - would integrate with Shopify Analytics API
        print("Note: Shopify Analytics integration requires additional setup")
        return pd.DataFrame(columns=['date', 'sessions', 'page_views', 'conversion_rate'])


# Removed - now using GoogleAnalytics4Connector from api_connectors


# Removed - now using InstagramConnector from api_connectors


# Removed - now using MetaAdsConnector and GoogleAdsConnector from api_connectors


# ============================================================================
# DATA AGGREGATION & CORRELATION
# ============================================================================

class MarketingAnalyticsFeed:
    """Main class to aggregate and correlate all marketing data"""
    
    def __init__(self, lookback_days: int = 365):
        self.lookback_days = lookback_days
        self.end_date = datetime.now()
        self.start_date = self.end_date - timedelta(days=lookback_days)
        
        # Load configuration
        self.config = self._load_config()
        
        # Initialize connectors
        self.shopify = ShopifyDataConnector(SECRETS_DIR / 'shopify_orders_credentials.env')
        self.shopify_analytics = ShopifyAnalyticsConnector(SECRETS_DIR / 'shopify_credentials.env')
        
        # Initialize ALL social media API connectors (will gracefully handle missing credentials)
        self.instagram = InstagramConnector(SECRETS_DIR / 'instagram_credentials.env')
        self.tiktok = TikTokConnector(SECRETS_DIR / 'tiktok_credentials.env')
        self.pinterest = PinterestConnector(SECRETS_DIR / 'pinterest_credentials.env')
        
        # Initialize Ad platform connectors
        self.meta_ads = MetaAdsConnector(SECRETS_DIR / 'meta_ads_credentials.env')
        self.google_ads = GoogleAdsConnector(
            SECRETS_DIR / 'google_ads_credentials.json',
            SECRETS_DIR / 'google_ads_config.env'
        )
        
        # Initialize analytics connectors
        self.google_analytics = GoogleAnalytics4Connector(
            Path(self.config.get('ga4_credentials_json', '')),
            self.config.get('ga4_property_id', '')
        )
        
        # Manual tracking files (fallback)
        self.manual_influencer_file = DATA_DIR / 'influencer_posts.json'
        self.manual_campaigns_file = DATA_DIR / 'ad_campaigns.json'
    
    def _load_config(self) -> Dict[str, Any]:
        """Load or create marketing configuration"""
        if CONFIG_FILE.exists():
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        else:
            # Create default config
            default_config = {
                "ga4_property_id": "",
                "ga4_credentials_json": str(SECRETS_DIR / "google_analytics_credentials.json"),
                "lookback_days": 365,
                "correlation_window_days": 7,
                "last_updated": datetime.now().isoformat()
            }
            with open(CONFIG_FILE, 'w') as f:
                json.dump(default_config, f, indent=2)
            print(f"Created default config at {CONFIG_FILE}")
            return default_config
    
    def generate_unified_feed(self) -> pd.DataFrame:
        """Generate unified timeline of all marketing activities and sales"""
        print(f"\n{'='*70}")
        print("MARKETING ANALYTICS FEED - Data Aggregation")
        print(f"{'='*70}")
        print(f"Date range: {self.start_date.date()} to {self.end_date.date()}")
        
        # Get all data sources
        print("\n1. Fetching Shopify sales data...")
        sales_data = self.shopify.get_daily_sales(self.start_date, self.end_date)
        print(f"   ✓ Loaded {len(sales_data)} days of sales data")
        
        print("\n2. Fetching website traffic data...")
        # Google Analytics for general web traffic
        ga_traffic_data = self.google_analytics.get_daily_traffic(self.start_date, self.end_date)
        # Shopify Analytics for store-specific traffic
        shopify_traffic_data = self.shopify_analytics.get_traffic_data(self.start_date, self.end_date)
        # Merge traffic sources
        traffic_data = self._merge_traffic_data(ga_traffic_data, shopify_traffic_data)
        print(f"   ✓ Loaded {len(traffic_data)} days of traffic data")
        
        print("\n3. Fetching social media data...")
        # Instagram - both owned posts and tagged posts
        instagram_posts = self.instagram.get_posts(self.start_date, self.end_date)
        tagged_posts = self.instagram.get_tagged_posts(self.start_date, self.end_date)
        
        # TikTok posts
        tiktok_posts = self.tiktok.get_posts(self.start_date, self.end_date)
        
        # Pinterest pins
        pinterest_pins = self.pinterest.get_pins(self.start_date, self.end_date)
        
        # Manual entries
        manual_posts = self._get_manual_influencer_posts(self.start_date, self.end_date)
        
        # Merge all social data
        influencer_data = pd.concat([instagram_posts, tagged_posts, tiktok_posts, pinterest_pins, manual_posts], ignore_index=True)
        influencer_data = influencer_data.drop_duplicates(subset=['date', 'post_url'], keep='first')
        print(f"   ✓ Loaded {len(influencer_data)} social media posts (Instagram + TikTok + Pinterest + Manual)")
        
        print("\n4. Fetching ad campaign data...")
        # Get Meta Ads campaigns
        meta_campaigns = self.meta_ads.get_campaign_events(self.start_date, self.end_date)
        meta_creatives = self.meta_ads.get_ad_creatives(self.start_date, self.end_date)
        
        # Get Google Ads campaigns
        google_campaigns = self.google_ads.get_campaign_events(self.start_date, self.end_date)
        
        # Get manual campaign entries
        manual_campaigns = self._get_manual_campaign_events(self.start_date, self.end_date)
        
        # Merge all campaign data
        campaign_data = pd.concat([meta_campaigns, meta_creatives, google_campaigns, manual_campaigns], ignore_index=True)
        campaign_data = campaign_data.drop_duplicates(subset=['date', 'campaign_name', 'event_type'], keep='first')
        print(f"   ✓ Loaded {len(campaign_data)} campaign events (automated + manual)")
        
        # Create unified timeline
        print("\n5. Building unified timeline...")
        timeline = self._build_timeline(sales_data, traffic_data, influencer_data, campaign_data)
        print(f"   ✓ Created timeline with {len(timeline)} days")
        
        return timeline
    
    def _build_timeline(self, sales_df: pd.DataFrame, traffic_df: pd.DataFrame,
                       influencer_df: pd.DataFrame, campaign_df: pd.DataFrame) -> pd.DataFrame:
        """Merge all data sources into a single timeline"""
        
        # Create date range
        date_range = pd.date_range(start=self.start_date, end=self.end_date, freq='D')
        timeline = pd.DataFrame({'date': date_range.date})
        
        # Merge sales data
        if not sales_df.empty:
            timeline = timeline.merge(sales_df, on='date', how='left')
        else:
            timeline['order_count'] = 0
            timeline['total_revenue'] = 0.0
            timeline['avg_order_value'] = 0.0
        
        # Merge traffic data
        if not traffic_df.empty:
            timeline = timeline.merge(traffic_df, on='date', how='left')
        else:
            timeline['sessions'] = 0
            timeline['users'] = 0
            timeline['page_views'] = 0
            timeline['avg_session_duration'] = 0.0
        
        # Add influencer events
        timeline['influencer_posts'] = 0
        timeline['influencer_details'] = ''
        if not influencer_df.empty:
            influencer_counts = influencer_df.groupby('date').size().reset_index(name='influencer_posts')
            influencer_details = influencer_df.groupby('date').apply(
                lambda x: '; '.join([f"{row['influencer']} ({row['platform']})" for _, row in x.iterrows()])
            ).reset_index(name='influencer_details')
            
            timeline = timeline.merge(influencer_counts, on='date', how='left', suffixes=('', '_new'))
            timeline = timeline.merge(influencer_details, on='date', how='left', suffixes=('', '_new'))
            timeline['influencer_posts'] = timeline['influencer_posts_new'].fillna(timeline['influencer_posts'])
            timeline['influencer_details'] = timeline['influencer_details_new'].fillna(timeline['influencer_details'])
            timeline = timeline.drop(columns=['influencer_posts_new', 'influencer_details_new'], errors='ignore')
        
        # Add campaign events
        timeline['campaign_events'] = 0
        timeline['campaign_details'] = ''
        if not campaign_df.empty:
            campaign_counts = campaign_df.groupby('date').size().reset_index(name='campaign_events')
            campaign_details = campaign_df.groupby('date').apply(
                lambda x: '; '.join([f"{row['campaign_name']} - {row['event_type']}" for _, row in x.iterrows()])
            ).reset_index(name='campaign_details')
            
            timeline = timeline.merge(campaign_counts, on='date', how='left', suffixes=('', '_new'))
            timeline = timeline.merge(campaign_details, on='date', how='left', suffixes=('', '_new'))
            timeline['campaign_events'] = timeline['campaign_events_new'].fillna(timeline['campaign_events'])
            timeline['campaign_details'] = timeline['campaign_details_new'].fillna(timeline['campaign_details'])
            timeline = timeline.drop(columns=['campaign_events_new', 'campaign_details_new'], errors='ignore')
        
        # Fill NaN values
        timeline = timeline.fillna(0)
        timeline['influencer_details'] = timeline['influencer_details'].replace(0, '')
        timeline['campaign_details'] = timeline['campaign_details'].replace(0, '')
        
        return timeline
    
    def _merge_traffic_data(self, ga_data: pd.DataFrame, shopify_data: pd.DataFrame) -> pd.DataFrame:
        """Merge Google Analytics and Shopify traffic data"""
        if ga_data.empty and shopify_data.empty:
            return pd.DataFrame(columns=['date', 'sessions', 'users', 'page_views', 'avg_session_duration', 
                                        'source_breakdown', 'shopify_sessions', 'top_countries'])
        
        if ga_data.empty:
            # Only Shopify data available
            shopify_data = shopify_data.rename(columns={'sessions': 'shopify_sessions'})
            shopify_data['sessions'] = shopify_data['shopify_sessions']
            shopify_data['users'] = shopify_data.get('visitors', 0)
            shopify_data['avg_session_duration'] = 0.0
            shopify_data['source_breakdown'] = shopify_data.get('top_sources', '')
            return shopify_data
        
        if shopify_data.empty:
            # Only GA data available
            ga_data['shopify_sessions'] = 0
            ga_data['top_countries'] = ''
            return ga_data
        
        # Both available - merge
        merged = ga_data.merge(
            shopify_data[['date', 'sessions', 'top_sources', 'top_countries']],
            on='date',
            how='outer',
            suffixes=('_ga', '_shopify')
        )
        
        # Combine source breakdown
        merged['source_breakdown'] = merged.apply(
            lambda row: f"GA: {row.get('source_breakdown', '')}; Shopify: {row.get('top_sources', '')}", axis=1
        )
        merged['shopify_sessions'] = merged.get('sessions_shopify', 0)
        merged['sessions'] = merged.get('sessions_ga', 0)  # Use GA as primary session count
        
        return merged.fillna(0)
    
    def _get_manual_influencer_posts(self, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """Get manually tracked influencer posts"""
        if self.manual_influencer_file.exists():
            with open(self.manual_influencer_file, 'r') as f:
                posts = json.load(f)
            
            df = pd.DataFrame(posts)
            if not df.empty:
                df['date'] = pd.to_datetime(df['date']).dt.date
                mask = (df['date'] >= start_date.date()) & (df['date'] <= end_date.date())
                return df[mask]
        
        return pd.DataFrame(columns=['date', 'platform', 'influencer', 'post_url', 'reach', 'engagement', 'notes'])
    
    def _get_manual_campaign_events(self, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """Get manually tracked campaign events"""
        if self.manual_campaigns_file.exists():
            with open(self.manual_campaigns_file, 'r') as f:
                events = json.load(f)
            
            df = pd.DataFrame(events)
            if not df.empty:
                df['date'] = pd.to_datetime(df['date']).dt.date
                mask = (df['date'] >= start_date.date()) & (df['date'] <= end_date.date())
                return df[mask]
        
        return pd.DataFrame(columns=['date', 'platform', 'campaign_name', 'event_type', 'budget', 'notes'])
    
    def add_manual_influencer_post(self, date: str, platform: str, influencer: str, 
                                  post_url: str, reach: int = 0, engagement: int = 0, notes: str = ""):
        """Manually add an influencer post"""
        posts = []
        if self.manual_influencer_file.exists():
            with open(self.manual_influencer_file, 'r') as f:
                posts = json.load(f)
        
        posts.append({
            'date': date,
            'platform': platform,
            'influencer': influencer,
            'post_url': post_url,
            'reach': reach,
            'engagement': engagement,
            'notes': notes,
            'added_at': datetime.now().isoformat()
        })
        
        with open(self.manual_influencer_file, 'w') as f:
            json.dump(posts, f, indent=2)
        
        print(f"✓ Added influencer post: {influencer} on {platform} ({date})")
    
    def add_manual_campaign_event(self, date: str, platform: str, campaign_name: str, 
                                 event_type: str, budget: float = 0, notes: str = ""):
        """Manually add a campaign event"""
        events = []
        if self.manual_campaigns_file.exists():
            with open(self.manual_campaigns_file, 'r') as f:
                events = json.load(f)
        
        events.append({
            'date': date,
            'platform': platform,
            'campaign_name': campaign_name,
            'event_type': event_type,
            'budget': budget,
            'notes': notes,
            'added_at': datetime.now().isoformat()
        })
        
        with open(self.manual_campaigns_file, 'w') as f:
            json.dump(events, f, indent=2)
        
        print(f"✓ Added campaign event: {campaign_name} - {event_type} ({date})")
    
    def calculate_correlations(self, timeline: pd.DataFrame) -> pd.DataFrame:
        """Calculate impact of marketing activities on sales and traffic"""
        window_days = self.config.get('correlation_window_days', 7)
        
        # Add rolling averages
        timeline['revenue_7day_avg'] = timeline['total_revenue'].rolling(window=7, min_periods=1).mean()
        timeline['traffic_7day_avg'] = timeline['sessions'].rolling(window=7, min_periods=1).mean()
        
        # Calculate day-over-day changes
        timeline['revenue_change_pct'] = timeline['total_revenue'].pct_change() * 100
        timeline['traffic_change_pct'] = timeline['sessions'].pct_change() * 100
        
        # Flag significant events
        timeline['has_marketing_event'] = (timeline['influencer_posts'] > 0) | (timeline['campaign_events'] > 0)
        
        # Calculate post-event impact (next N days after an event)
        timeline['post_event_revenue_avg'] = 0.0
        timeline['post_event_traffic_avg'] = 0.0
        
        for idx in timeline[timeline['has_marketing_event']].index:
            # Get next N days
            end_idx = min(idx + window_days, len(timeline))
            if end_idx > idx:
                timeline.loc[idx, 'post_event_revenue_avg'] = timeline.loc[idx:end_idx-1, 'total_revenue'].mean()
                timeline.loc[idx, 'post_event_traffic_avg'] = timeline.loc[idx:end_idx-1, 'sessions'].mean()
        
        return timeline
    
    def _collect_all_social_posts(self) -> pd.DataFrame:
        """Collect all social media posts from all platforms for detailed view"""
        all_posts = []
        
        # Instagram
        instagram_posts = self.instagram.get_posts(self.start_date, self.end_date)
        if not instagram_posts.empty:
            all_posts.append(instagram_posts)
        
        # TikTok
        tiktok_posts = self.tiktok.get_posts(self.start_date, self.end_date)
        if not tiktok_posts.empty:
            all_posts.append(tiktok_posts)
        
        # Pinterest
        pinterest_pins = self.pinterest.get_pins(self.start_date, self.end_date)
        if not pinterest_pins.empty:
            all_posts.append(pinterest_pins)
        
        # Manual posts
        manual_posts = self._get_manual_influencer_posts(self.start_date, self.end_date)
        if not manual_posts.empty:
            all_posts.append(manual_posts)
        
        if all_posts:
            combined = pd.concat(all_posts, ignore_index=True)
            combined = combined.sort_values('date', ascending=False)
            return combined
        
        return pd.DataFrame()
    
    def _create_social_media_sheet(self, worksheet, posts_df: pd.DataFrame, timeline_df: pd.DataFrame):
        """Create detailed social media posts sheet with impact metrics"""
        if posts_df.empty:
            worksheet.cell(1, 1, "No social media data available")
            return
        
        # Add impact metrics to each post
        posts_with_impact = posts_df.copy()
        posts_with_impact['7_day_revenue_impact'] = 0.0
        posts_with_impact['7_day_traffic_impact'] = 0.0
        
        for idx, post in posts_with_impact.iterrows():
            post_date = post['date']
            # Find corresponding timeline data
            timeline_match = timeline_df[timeline_df['date'] == post_date]
            if not timeline_match.empty:
                posts_with_impact.loc[idx, '7_day_revenue_impact'] = timeline_match.iloc[0].get('post_event_revenue_avg', 0)
                posts_with_impact.loc[idx, '7_day_traffic_impact'] = timeline_match.iloc[0].get('post_event_traffic_avg', 0)
        
        # Write header
        headers = ['Date', 'Platform', 'Influencer/Account', 'Post URL', 'Views', 'Reach', 'Impressions', 
                  'Likes', 'Comments', 'Shares', 'Saves', 'Total Engagement', 
                  '7-Day Revenue Impact', '7-Day Traffic Impact', 'Notes']
        
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Write data
        for row_idx, (_, post) in enumerate(posts_with_impact.iterrows(), 2):
            worksheet.cell(row_idx, 1, post['date'])
            worksheet.cell(row_idx, 2, post.get('platform', ''))
            worksheet.cell(row_idx, 3, post.get('influencer', ''))
            worksheet.cell(row_idx, 4, post.get('post_url', ''))
            worksheet.cell(row_idx, 5, post.get('views', 0))
            worksheet.cell(row_idx, 6, post.get('reach', 0))
            worksheet.cell(row_idx, 7, post.get('impressions', 0))
            worksheet.cell(row_idx, 8, post.get('likes', 0))
            worksheet.cell(row_idx, 9, post.get('comments', 0))
            worksheet.cell(row_idx, 10, post.get('shares', 0))
            worksheet.cell(row_idx, 11, post.get('saves', 0))
            worksheet.cell(row_idx, 12, post.get('engagement', 0))
            worksheet.cell(row_idx, 13, f"${post.get('7_day_revenue_impact', 0):,.2f}")
            worksheet.cell(row_idx, 14, int(post.get('7_day_traffic_impact', 0)))
            worksheet.cell(row_idx, 15, post.get('notes', ''))
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 40
        worksheet.column_dimensions['M'].width = 18
        worksheet.column_dimensions['N'].width = 18
        worksheet.column_dimensions['O'].width = 50
    
    def _create_heatmap_sheet(self, worksheet, timeline_df: pd.DataFrame):
        """Create a heat map view showing post impact by day/week"""
        # Group by week
        timeline_df['week'] = pd.to_datetime(timeline_df['date']).dt.isocalendar().week
        timeline_df['year'] = pd.to_datetime(timeline_df['date']).dt.year
        timeline_df['day_name'] = pd.to_datetime(timeline_df['date']).dt.day_name()
        
        # Aggregate by week
        weekly = timeline_df.groupby(['year', 'week']).agg({
            'total_revenue': 'sum',
            'sessions': 'sum',
            'influencer_posts': 'sum',
            'campaign_events': 'sum'
        }).reset_index()
        
        # Write header
        headers = ['Year', 'Week', 'Total Revenue', 'Total Sessions', 'Influencer Posts', 'Campaign Events', 'Marketing Activity Level']
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data with conditional formatting (heat map colors)
        max_revenue = weekly['total_revenue'].max() if not weekly.empty else 1
        max_traffic = weekly['sessions'].max() if not weekly.empty else 1
        
        for row_idx, (_, week_data) in enumerate(weekly.iterrows(), 2):
            worksheet.cell(row_idx, 1, week_data['year'])
            worksheet.cell(row_idx, 2, week_data['week'])
            worksheet.cell(row_idx, 3, f"${week_data['total_revenue']:,.2f}")
            worksheet.cell(row_idx, 4, int(week_data['sessions']))
            worksheet.cell(row_idx, 5, int(week_data['influencer_posts']))
            worksheet.cell(row_idx, 6, int(week_data['campaign_events']))
            
            # Marketing activity level
            activity_level = "High" if (week_data['influencer_posts'] + week_data['campaign_events']) >= 3 else \
                           "Medium" if (week_data['influencer_posts'] + week_data['campaign_events']) >= 1 else "Low"
            cell = worksheet.cell(row_idx, 7, activity_level)
            
            # Color code based on activity
            if activity_level == "High":
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif activity_level == "Medium":
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    
    def _create_traffic_sources_sheet(self, worksheet, timeline_df: pd.DataFrame):
        """Create traffic sources breakdown visualization"""
        # Extract top sources from source_breakdown column
        worksheet.cell(1, 1, "Traffic Sources Analysis")
        worksheet.cell(1, 1).font = Font(bold=True, size=14)
        
        worksheet.cell(3, 1, "Date")
        worksheet.cell(3, 2, "Total Sessions")
        worksheet.cell(3, 3, "Source Breakdown")
        
        for col in range(1, 4):
            worksheet.cell(3, col).font = Font(bold=True)
            worksheet.cell(3, col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            worksheet.cell(3, col).font = Font(bold=True, color="FFFFFF")
        
        # Write traffic data
        for row_idx, (_, day) in enumerate(timeline_df.iterrows(), 4):
            worksheet.cell(row_idx, 1, day['date'])
            worksheet.cell(row_idx, 2, int(day.get('sessions', 0)))
            worksheet.cell(row_idx, 3, day.get('source_breakdown', ''))
        
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 80
    
    def _create_geography_sheet(self, worksheet, timeline_df: pd.DataFrame):
        """Create geographic breakdown of Shopify traffic"""
        worksheet.cell(1, 1, "Geographic Traffic Analysis (Shopify)")
        worksheet.cell(1, 1).font = Font(bold=True, size=14)
        
        worksheet.cell(3, 1, "Date")
        worksheet.cell(3, 2, "Shopify Sessions")
        worksheet.cell(3, 3, "Top Countries")
        
        for col in range(1, 4):
            worksheet.cell(3, col).font = Font(bold=True)
            worksheet.cell(3, col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            worksheet.cell(3, col).font = Font(bold=True, color="FFFFFF")
        
        # Write geographic data
        for row_idx, (_, day) in enumerate(timeline_df.iterrows(), 4):
            worksheet.cell(row_idx, 1, day['date'])
            worksheet.cell(row_idx, 2, int(day.get('shopify_sessions', 0)))
            worksheet.cell(row_idx, 3, day.get('top_countries', ''))
        
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 60
    
    def export_to_excel(self, timeline: pd.DataFrame):
        """Export unified feed to Excel with ENHANCED formatting and visualizations"""
        print(f"\n6. Exporting to Excel: {OUTPUT_FILE}")
        
        # Calculate correlations
        timeline_with_analysis = self.calculate_correlations(timeline)
        
        # Collect detailed social media posts for separate sheet
        all_social_posts = self._collect_all_social_posts()
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Timeline (all data)
        ws_timeline = wb.active
        ws_timeline.title = "Timeline"
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(timeline_with_analysis, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_timeline.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Highlight marketing event rows
                if r_idx > 1 and timeline_with_analysis.iloc[r_idx-2]['has_marketing_event']:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Adjust column widths
        ws_timeline.column_dimensions['A'].width = 12  # date
        ws_timeline.column_dimensions['B'].width = 12  # order_count
        ws_timeline.column_dimensions['C'].width = 15  # total_revenue
        ws_timeline.column_dimensions['D'].width = 15  # avg_order_value
        ws_timeline.column_dimensions['E'].width = 12  # sessions
        ws_timeline.column_dimensions['F'].width = 12  # users
        
        # Sheet 2: Detailed Social Media Posts
        ws_social = wb.create_sheet("Social Media Details")
        self._create_social_media_sheet(ws_social, all_social_posts, timeline_with_analysis)
        
        # Sheet 3: Marketing Events Summary
        ws_events = wb.create_sheet("Marketing Events")
        events_data = timeline_with_analysis[timeline_with_analysis['has_marketing_event']].copy()
        
        if not events_data.empty:
            for r_idx, row in enumerate(dataframe_to_rows(events_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_events.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Sheet 4: Heat Map / Impact Analysis
        ws_heatmap = wb.create_sheet("Heat Map")
        self._create_heatmap_sheet(ws_heatmap, timeline_with_analysis)
        
        # Sheet 5: Traffic Sources Breakdown
        ws_traffic = wb.create_sheet("Traffic Sources")
        self._create_traffic_sources_sheet(ws_traffic, timeline_with_analysis)
        
        # Sheet 6: Geographic Analysis
        ws_geo = wb.create_sheet("Geography")
        self._create_geography_sheet(ws_geo, timeline_with_analysis)
        
        # Sheet 7: Summary Stats
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Date Range", f"{self.start_date.date()} to {self.end_date.date()}"],
            ["Total Days", len(timeline)],
            ["Total Revenue", f"${timeline['total_revenue'].sum():,.2f}"],
            ["Average Daily Revenue", f"${timeline['total_revenue'].mean():,.2f}"],
            ["Total Orders", int(timeline['order_count'].sum())],
            ["Total Sessions", int(timeline['sessions'].sum())],
            ["Influencer Posts", int(timeline['influencer_posts'].sum())],
            ["Campaign Events", int(timeline['campaign_events'].sum())],
            ["Days with Marketing Activity", int(timeline['has_marketing_event'].sum())],
        ]
        
        for r_idx, row in enumerate(summary_data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        
        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"   ✓ Saved to {OUTPUT_FILE}")
        print(f"\n{'='*70}")
        print("✓ Marketing Analytics Feed Generated Successfully!")
        print(f"{'='*70}\n")


# ============================================================================
# CLI INTERFACE
# ============================================================================

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Marketing Analytics Feed Generator')
    parser.add_argument('--days', type=int, default=365, help='Number of days to look back (default: 365)')
    parser.add_argument('--add-influencer', action='store_true', help='Add an influencer post')
    parser.add_argument('--add-campaign', action='store_true', help='Add a campaign event')
    
    args = parser.parse_args()
    
    # Initialize feed
    feed = MarketingAnalyticsFeed(lookback_days=args.days)
    
    # Add influencer post manually
    if args.add_influencer:
        print("\n=== Add Influencer Post ===")
        date = input("Date (YYYY-MM-DD): ")
        platform = input("Platform (Instagram/TikTok/YouTube/etc): ")
        influencer = input("Influencer name: ")
        post_url = input("Post URL: ")
        reach = int(input("Reach (0 if unknown): ") or "0")
        engagement = int(input("Engagement (0 if unknown): ") or "0")
        notes = input("Notes (optional): ")
        
        feed.add_manual_influencer_post(date, platform, influencer, post_url, reach, engagement, notes)
        return
    
    # Add campaign event manually
    if args.add_campaign:
        print("\n=== Add Campaign Event ===")
        date = input("Date (YYYY-MM-DD): ")
        platform = input("Platform (Meta/Google/TikTok/etc): ")
        campaign_name = input("Campaign name: ")
        event_type = input("Event type (launch/creative_change/budget_change/pause/resume): ")
        budget = float(input("Budget (0 if not applicable): ") or "0")
        notes = input("Notes (optional): ")
        
        feed.add_manual_campaign_event(date, platform, campaign_name, event_type, budget, notes)
        return
    
    # Generate feed
    timeline = feed.generate_unified_feed()
    feed.export_to_excel(timeline)


if __name__ == '__main__':
    main()
